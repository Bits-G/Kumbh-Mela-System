# from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
# from django.contrib.auth.decorators import login_required
# from django.contrib import messages
# from django.shortcuts import render, redirect, get_object_or_404
# from django.db.models import Q
# from django.http import HttpResponse
# from django.core.paginator import Paginator
# from django.template.loader import render_to_string

# from .models import Officer
# from .forms import OfficerForm


# def admin_login(request):
#     if request.user.is_authenticated:
#         return redirect('kumbh_mela_dashboard')

#     if request.method == 'POST':
#         username = request.POST.get('username')
#         password = request.POST.get('password')
#         user = authenticate(request, username=username, password=password)
#         if user is not None:
#             auth_login(request, user)
#             return redirect('kumbh_mela_dashboard')
#         else:
#             messages.error(request, 'गलत यूज़रनेम या पासवर्ड (Invalid credentials)')

#     return render(request, 'officers/login.html')


# @login_required
# def admin_logout(request):
#     auth_logout(request)
#     return redirect('admin_login')


# @login_required
# def kumbh_mela_dashboard(request):
#     total_officers = Officer.objects.count()
#     return render(request, 'officers/kumbh_mela.html', {'total_officers': total_officers})


# @login_required
# def officer_list(request):
#     officers = Officer.objects.all()
#     paginator = Paginator(officers, 25)
#     page_number = request.GET.get('page')
#     page_obj = paginator.get_page(page_number)
#     return render(request, 'officers/officer_list.html', {'page_obj': page_obj})


# @login_required
# def officer_add(request):
#     if request.method == 'POST':
#         form = OfficerForm(request.POST)
#         if form.is_valid():
#             form.save()
#             messages.success(request, 'Entry सफलतापूर्वक जोड़ी गई (Added successfully)')
#             return redirect('officer_list')
#     else:
#         form = OfficerForm()
#     return render(request, 'officers/officer_form.html', {'form': form, 'action': 'Add'})


# @login_required
# def officer_edit(request, pk):
#     officer = get_object_or_404(Officer, pk=pk)
#     if request.method == 'POST':
#         form = OfficerForm(request.POST, instance=officer)
#         if form.is_valid():
#             form.save()
#             messages.success(request, 'Entry सफलतापूर्वक अपडेट हुई (Updated successfully)')
#             return redirect('officer_list')
#     else:
#         form = OfficerForm(instance=officer)
#     return render(request, 'officers/officer_form.html', {'form': form, 'action': 'Edit'})


# @login_required
# def officer_delete(request, pk):
#     officer = get_object_or_404(Officer, pk=pk)
#     if request.method == 'POST':
#         officer.delete()
#         messages.success(request, 'Entry हटा दी गई (Deleted successfully)')
#         return redirect('officer_list')
#     return render(request, 'officers/officer_confirm_delete.html', {'officer': officer})


# @login_required
# def search_officers(request):
#     query = request.GET.get('q', '').strip()
#     search_by = request.GET.get('search_by', 'name')

#     field_map = {
#         'name': 'name__icontains',
#         'mobile': 'mobile_number__icontains',
#         'address': 'address__icontains',
#         'designation': 'designation__icontains',
#         'city': 'city__icontains',
#         'state': 'state__icontains',
#     }

#     results = []
#     if query and search_by in field_map:
#         results = Officer.objects.filter(**{field_map[search_by]: query})[:200]

#     return render(request, 'officers/search_results.html', {
#         'results': results,
#         'query': query,
#         'search_by': search_by,
#     })


# @login_required
# def generate_officer_pdf(request, pk):
#     officer = get_object_or_404(Officer, pk=pk)
#     html_string = render_to_string('officers/officer_pdf_template.html', {'officer': officer})

#     try:
#         from weasyprint import HTML
#         pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()
#         response = HttpResponse(pdf_file, content_type='application/pdf')
#         response['Content-Disposition'] = f'inline; filename="officer_{officer.id}.pdf"'
#         return response
#     except Exception as e:
#         return HttpResponse(
#             f"PDF generation library not fully installed on this server. Error: {e}.",
#             status=500
#         )










import csv

import io
import qrcode
import base64
import json
from io import BytesIO
from urllib.parse import urlencode
from openpyxl import Workbook

from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Q
from django.http import HttpResponse
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.template.loader import render_to_string
from django.urls import reverse
from openpyxl import load_workbook
from .models import SUB_TYPE_MAP, SUB_SUB_TYPE_MAP
from .models import Officer
from .forms import OfficerForm
from django.shortcuts import redirect



def admin_login(request):
    if request.user.is_authenticated:
        return redirect('kumbh_mela_dashboard')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None:
            auth_login(request, user)
            return redirect('kumbh_mela_dashboard')
        else:
            messages.error(request, 'गलत यूज़रनेम या पासवर्ड (Invalid credentials)')
    return render(request, 'officers/login.html')





@login_required
def admin_logout(request):
    auth_logout(request)
    return redirect('admin_login')





# @login_required

# def kumbh_mela_dashboard(request):

#     total_officers = Officer.objects.count()

#     return render(request, 'officers/kumbh_mela.html', {'total_officers': total_officers})


# 17 july 10:26pm
# @login_required
# def kumbh_mela_dashboard(request):
#     from .models import DIVISION_CHOICES, GOVT_LEVEL_CHOICES

#     total_officers = Officer.objects.count()

#     govt_level_counts = []
#     for value, label in GOVT_LEVEL_CHOICES:
#         count = Officer.objects.filter(government_level=value).count()
#         govt_level_counts.append({'value': value, 'label': label, 'count': count})

#     division_counts = []
#     for value, label in DIVISION_CHOICES:
#         count = Officer.objects.filter(division=value).count()
#         division_counts.append({'value': value, 'label': label, 'count': count})

#     return render(request, 'officers/kumbh_mela.html', {
#         'total_officers': total_officers,
#         'govt_level_counts': govt_level_counts,
#         'division_counts': division_counts,
#     })
@login_required
def kumbh_mela_dashboard(request):
    from .models import DIVISION_CHOICES, GOVT_LEVEL_CHOICES, MAIN_TYPE_CHOICES, SUB_TYPE_MAP, SUB_SUB_TYPE_MAP

    total_officers = Officer.objects.count()

    govt_level_counts = []
    for value, label in GOVT_LEVEL_CHOICES:
        count = Officer.objects.filter(government_level=value).count()
        govt_level_counts.append({'value': value, 'label': label, 'count': count})

    division_counts = []
    for value, label in DIVISION_CHOICES:
        count = Officer.objects.filter(division=value).count()
        division_counts.append({'value': value, 'label': label, 'count': count})

    # Main Type cards (always shown)
    main_type_counts = []
    for value, label in MAIN_TYPE_CHOICES:
        count = Officer.objects.filter(main_type=value).count()
        main_type_counts.append({'value': value, 'label': label, 'count': count})

    selected_main = request.GET.get('main')
    selected_sub = request.GET.get('sub')

    # Sub Type cards (shown if a Main Type is selected)
    sub_type_counts = []
    if selected_main and selected_main in SUB_TYPE_MAP:
        for value, label in SUB_TYPE_MAP[selected_main]:
            count = Officer.objects.filter(main_type=selected_main, sub_type=value).count()
            sub_type_counts.append({'value': value, 'label': label, 'count': count})

    # Sub-Sub Type cards (shown if a Sub Type is selected)
    sub_sub_type_counts = []
    if selected_sub and selected_sub in SUB_SUB_TYPE_MAP:
        for value, label in SUB_SUB_TYPE_MAP[selected_sub]:
            count = Officer.objects.filter(sub_type=selected_sub, sub_sub_type=value).count()
            sub_sub_type_counts.append({'value': value, 'label': label, 'count': count})

    matching_entries = []
    if selected_main:
        entries_qs = Officer.objects.filter(main_type=selected_main)
        if selected_sub:
            entries_qs = entries_qs.filter(sub_type=selected_sub)
        matching_entries = entries_qs[:15]

    return render(request, 'officers/kumbh_mela.html', {
        'total_officers': total_officers,
        'govt_level_counts': govt_level_counts,
        'division_counts': division_counts,
        'main_type_counts': main_type_counts,
        'sub_type_counts': sub_type_counts,
        'sub_sub_type_counts': sub_sub_type_counts,
        'selected_main': selected_main,
        'selected_sub': selected_sub,
        'matching_entries': matching_entries,
    })



# @login_required

# def officer_list(request):

#     officers = Officer.objects.all()

#     paginator = Paginator(officers, 25)

#     page_number = request.GET.get('page')

#     page_obj = paginator.get_page(page_number)

#     return render(request, 'officers/officer_list.html', {'page_obj': page_obj})

@login_required
def officer_list(request):
    officers = Officer.objects.all()

    division_filter = request.GET.get('division')
    govt_level_filter = request.GET.get('govt_level')
    main_type_filter = request.GET.get('main_type')
    sub_type_filter = request.GET.get('sub_type')
    sub_sub_type_filter = request.GET.get('sub_sub_type')
    

    if division_filter:
        officers = officers.filter(division=division_filter)
    if govt_level_filter:
        officers = officers.filter(government_level=govt_level_filter)
    if main_type_filter:
        officers = officers.filter(main_type=main_type_filter)
    if sub_type_filter:
        officers = officers.filter(sub_type=sub_type_filter)
    if sub_sub_type_filter:
        officers = officers.filter(sub_sub_type=sub_sub_type_filter)

    paginator = Paginator(officers, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'officers/officer_list.html', {
        'page_obj': page_obj,
        'division_filter': division_filter,
        'govt_level_filter': govt_level_filter,
    })




@login_required
def officer_add(request):
    just_added = False
    if request.method == 'POST':
        form = OfficerForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            just_added = True
            form = OfficerForm()  # blank form ready for next entry
    else:
        form = OfficerForm()

    return render(request, 'officers/officer_form.html', {
        'form': form, 'action': 'Add', 'just_added': just_added,
        'sub_type_map_json': json.dumps(SUB_TYPE_MAP),
        'sub_sub_type_map_json': json.dumps(SUB_SUB_TYPE_MAP),
        })





@login_required
def officer_edit(request, pk):
    officer = get_object_or_404(Officer, pk=pk)
    if request.method == 'POST':
        form = OfficerForm(request.POST, request.FILES, instance=officer)
        if form.is_valid():
            form.save()
            messages.success(request, 'Entry सफलतापूर्वक अपडेट हुई (Updated successfully)')
            return redirect('officer_list')
    else:
        form = OfficerForm(instance=officer)

    return render(request, 'officers/officer_form.html', {
        'form': form, 'action': 'Edit',
        'sub_type_map_json': json.dumps(SUB_TYPE_MAP),
        'sub_sub_type_map_json': json.dumps(SUB_SUB_TYPE_MAP),
        })





@login_required
def officer_delete(request, pk):
    officer = get_object_or_404(Officer, pk=pk)
    if request.method == 'POST':
        officer.delete()
        messages.success(request, 'Entry हटा दी गई (Deleted successfully)')
        return redirect('officer_list')
    return render(request, 'officers/officer_confirm_delete.html', {'officer': officer})





@login_required
def officer_import(request):
    imported_count = 0
    error_message = None

    if request.method == 'POST' and request.FILES.get('data_file'):
        uploaded_file = request.FILES['data_file']
        filename = uploaded_file.name.lower()

        try:
            rows = []
            if filename.endswith('.csv'):
                decoded = uploaded_file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(decoded))
                for row in reader:
                    rows.append(row)
            elif filename.endswith('.xlsx'):
                wb = load_workbook(uploaded_file, read_only=True)
                ws = wb.active
                headers = [str(cell.value).strip() if cell.value else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                for excel_row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, excel_row)))
            else:
                error_message = 'Sirf .csv ya .xlsx file allowed hai.'

            if not error_message:
                batch = []
                for row in rows:
                    name = (row.get('name') or '').strip()
                    if not name:
                        continue
                    batch.append(Officer(
                        name=name,
                        mobile_number=str(row.get('mobile_number') or '').strip(),
                        designation=(row.get('designation') or '').strip(),                        
                        state=(row.get('state') or '').strip(),                        
                        city=(row.get('city') or '').strip(),
                        address=(row.get('address') or '').strip(),
                        work=(row.get('work') or '').strip(),
                    ))
                Officer.objects.bulk_create(batch, batch_size=1000)
                imported_count = len(batch)
        except Exception as e:
            error_message = f'File process karte waqt error aaya: {e}'

    return render(request, 'officers/officer_import.html', {
        'imported_count': imported_count,
        'error_message': error_message,
    })





@login_required
def search_officers(request):
    query = request.GET.get('q', '').strip()
    search_by = request.GET.get('search_by', 'name')
    field_map = {
        'name': 'name__icontains',
        'mobile': 'mobile_number__icontains',
        'address': 'address__icontains',
        'designation': 'designation__icontains',
        'department': 'department__icontains',
        'category': 'category__icontains',
        'sub_category': 'sub_category__icontains',
        'contact_2': 'contact_2__icontains',
        'email': 'email__icontains',
        'office_phone': 'office_phone__icontains',
    }

    results = []
    pdf_full_url = None
    if query and search_by in field_map:
        results = Officer.objects.filter(**{field_map[search_by]: query})[:200]
        pdf_relative_url = reverse('search_results_pdf') + '?' + urlencode({'q': query, 'search_by': search_by})
        pdf_full_url = request.build_absolute_uri(pdf_relative_url)

    return render(request, 'officers/search_results.html', {
        'results': results,
        'query': query,
        'search_by': search_by,
        'pdf_full_url': pdf_full_url,
    })





@login_required
def search_results_pdf(request):
    query = request.GET.get('q', '').strip()
    search_by = request.GET.get('search_by', 'name')
    field_map = {
        'name': 'name__icontains',
        'mobile': 'mobile_number__icontains',
        'address': 'address__icontains',
        'designation': 'designation__icontains',
        'department': 'department__icontains',
        'category': 'category__icontains',
        'sub_category': 'sub_category__icontains',
        'contact_2': 'contact_2__icontains',
        'email': 'email__icontains',
        'office_phone': 'office_phone__icontains',
    }

    results = []
    if query and search_by in field_map:
        results = Officer.objects.filter(**{field_map[search_by]: query})
    html_string = render_to_string('officers/officer_search_pdf_template.html', {
        'results': results,
        'query': query,
    })
    try:
        from weasyprint import HTML
        pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="search_results.pdf"'
        return response
    except Exception as e:
        return HttpResponse(f"PDF generation error: {e}", status=500)





# @login_required
# def generate_officer_pdf(request, pk):
#     officer = get_object_or_404(Officer, pk=pk)
#     html_string = render_to_string('officers/officer_pdf_template.html', {'officer': officer})

#     try:
#         from weasyprint import HTML
#         pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()
#         response = HttpResponse(pdf_file, content_type='application/pdf')
#         response['Content-Disposition'] = f'inline; filename="officer_{officer.id}.pdf"'
#         return response
#     except Exception as e:
#         return HttpResponse(
#             f"PDF generation library not fully installed on this server. Error: {e}.",
#             status=500
#         )


# new implementation of generate_officer_pdf with QR code generation
@login_required
def generate_officer_pdf(request, pk):
    officer = get_object_or_404(Officer, pk=pk)

    pdf_view_url = request.build_absolute_uri(reverse('generate_officer_pdf', args=[officer.pk]))
    qr = qrcode.make(pdf_view_url)
    buffer = BytesIO()
    qr.save(buffer, format='PNG')
    qr_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')

    html_string = render_to_string('officers/officer_pdf_template.html', {
        'officer': officer,
        'qr_base64': qr_base64,
    })

    try:
        from weasyprint import HTML
        pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="officer_{officer.id}.pdf"'
        return response
    except Exception as e:
        return HttpResponse(
            f"PDF generation library not fully installed on this server. Error: {e}.",
            status=500
        )
    
@login_required
def officer_qr_code(request, pk):
    officer = get_object_or_404(Officer, pk=pk)
    pdf_url = request.build_absolute_uri(reverse('generate_officer_pdf', args=[officer.pk]))

    qr = qrcode.make(pdf_url)
    buffer = BytesIO()
    qr.save(buffer, format='PNG')

    return HttpResponse(buffer.getvalue(), content_type='image/png')


@login_required
def search_results_qr(request):
    query = request.GET.get('q', '').strip()
    search_by = request.GET.get('search_by', 'name')
    pdf_relative_url = reverse('search_results_pdf') + '?' + urlencode({'q': query, 'search_by': search_by})
    pdf_full_url = request.build_absolute_uri(pdf_relative_url)
    qr = qrcode.make(pdf_full_url)
    buffer = BytesIO()
    qr.save(buffer, format='PNG')

    return HttpResponse(buffer.getvalue(), content_type='image/png')




def officer_export_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="officers.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'name',
        'mobile_number',
        'designation',
        'city',
        'state',
        'address',
        'work'
    ])

    for officer in Officer.objects.all():
        writer.writerow([
            officer.name,
            officer.mobile_number,
            officer.designation,
            officer.city,
            officer.state,
            officer.address,
            officer.work
        ])

    return response

def officer_export(request):
    return render(request, 'officers/officer_export.html')

def officer_export_excel(request):

    wb = Workbook()
    ws = wb.active
    ws.title = "Officers"
    ws.append([
        'name',
        'mobile_number',
        'designation',
        'city',
        'state',
        'address',
        'work'
    ])

    for officer in Officer.objects.all():
        ws.append([
            officer.name,
            officer.mobile_number,
            officer.designation,
            officer.city,
            officer.state,
            officer.address,
            officer.work
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=officers.xlsx'
    wb.save(response)
    return response

@login_required
def delete_selected_officers(request):
    if request.method == "POST":
        ids = request.POST.getlist("selected_ids")
        if ids:
            Officer.objects.filter(id__in=ids).delete()
            messages.success(
                request,
                f"{len(ids)} Entries deleted successfully."
            )
        else:
            messages.warning(
                request,
                "Please select at least one entry."
            )
    return redirect("officer_list")

@login_required
def selected_officers_pdf(request):
    ids_param = request.GET.get('ids', '')
    ids = [i for i in ids_param.split(',') if i.isdigit()]
    results = Officer.objects.filter(pk__in=ids)

    html_string = render_to_string('officers/officer_search_pdf_template.html', {
        'results': results,
        'query': 'Selected Entries',
    })

    try:
        from weasyprint import HTML
        pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="selected_entries.pdf"'
        return response
    except Exception as e:
        return HttpResponse(f"PDF generation error: {e}", status=500)
    

@login_required
def export_selected_csv(request):
    ids_param = request.GET.get('ids', '')
    ids = [i for i in ids_param.split(',') if i.isdigit()]
    officers = Officer.objects.filter(pk__in=ids)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="selected_entries.csv"'

    writer = csv.writer(response)
    writer.writerow(['Sr.No.', 'Name', 'Dept.', 'Designation', 'Category', 'Sub-Category',
                      'Contact 1', 'Contact 2', 'Email', 'Office Ph.', 'PBX Ext.'])

    for idx, officer in enumerate(officers, start=1):
        writer.writerow([
            idx, officer.name, officer.department or '', officer.designation or '',
            officer.category or '', officer.sub_category or '', officer.mobile_number,
            officer.contact_2 or '', officer.email or '', officer.office_phone or '',
            officer.pbx_extension or '',
        ])

    return response


@login_required
def export_selected_excel(request):
    from openpyxl import Workbook

    ids_param = request.GET.get('ids', '')
    ids = [i for i in ids_param.split(',') if i.isdigit()]
    officers = Officer.objects.filter(pk__in=ids)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Selected Entries'
    ws.append(['Sr.No.', 'Name', 'Dept.', 'Designation', 'Category', 'Sub-Category',
               'Contact 1', 'Contact 2', 'Email', 'Office Ph.', 'PBX Ext.'])

    for idx, officer in enumerate(officers, start=1):
        ws.append([
            idx, officer.name, officer.department or '', officer.designation or '',
            officer.category or '', officer.sub_category or '', officer.mobile_number,
            officer.contact_2 or '', officer.email or '', officer.office_phone or '',
            officer.pbx_extension or '',
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="selected_entries.xlsx"'
    wb.save(response)
    return response